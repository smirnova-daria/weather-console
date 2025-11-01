import datetime
from os import path

import requests
from openpyxl import Workbook, load_workbook

API_KEY = 'fd85cfdc7442ada7956eea08fb30064a'
API_URL = 'https://api.openweathermap.org/data/2.5/weather'
UNITS = 'metric'
LANG = 'ru'
FILE_EXCEL = 'weather_data.xlsx'


def get_weather(city):
    params = {
        'appid': API_KEY,
        'units': UNITS,
        'lang': LANG,
        'q': city,
    }
    try:
        r = requests.get(API_URL, params=params)
        return r.json()
    except:
        return {'cod': 0, 'message': 'Не удалось получить данные'}


def wind_direction(degrees):
    degrees = degrees % 360

    directions = [
        "Северный", "Северо-Восточный", "Восточный", "Юго-Восточный",
        "Южный", "Юго-Западный", "Западный", "Северо-Западный"
    ]

    index = round(degrees / 45) % 8

    return directions[index]


def pretty_time(timestamp, timezone_offset):
    tz = datetime.timezone(datetime.timedelta(seconds=timezone_offset))
    dt = datetime.datetime.fromtimestamp(timestamp, tz)
    return dt.strftime("%H:%M")


def get_weather_desc(data):
    if data['cod'] != 200:
        return data['message'] if 'message' in data else 'Ошибка получения данных :('

    desc = f"""Погода в г. {data['name']}
{data['weather'][0]['description'].capitalize()}
Температура {data['main']['temp']}°C 🌡️
Ощущается как {data['main']['feels_like']}°C
Ветер {wind_direction(data['wind']['deg'])}, {data['wind']['speed']} м/с 🎐
Давление {round(data['main']['pressure'] * 100 // 133.322)} мм.рт.ст
Влажность {data['main']['humidity']}% 💧
Восход солнца в {pretty_time(data['sys']['sunrise'], data['timezone'])} 🌅
Закат солнца в {pretty_time(data['sys']['sunset'], data['timezone'])} 🌇
    """

    return desc


def save_excel(data):
    if data['cod'] == 200:
        if path.exists(FILE_EXCEL):
            wb = load_workbook(FILE_EXCEL)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Статистика запросов'
            ws.append([
                'Дата запроса',
                'Город',
                'Температура, °C',
                'Ветер, м/с',
                'Давление, мм.рт.ст',
                'Влажность, %',
            ])

        ws.append([
            datetime.datetime.now(),
            data['name'],
            data['main']['temp'],
            f'{wind_direction(data['wind']['deg'])}, {data['wind']['speed']}',
            round(data['main']['pressure'] * 100 // 133.322),
            data['main']['humidity'],
        ])
        wb.save(filename=FILE_EXCEL)


print('*' * 70)
print("""o(*￣▽￣*)ブ Привет
Чтобы получить прогноз погоды, напиши название города
Чтобы выйти, нажми Enter
🪄🌤️🌧️❄️🌈""")
print('*' * 70)

while True:
    q = input("Город: ")
    if not q:
        print("👋Буду ждать тебя снова")
        break
    else:
        weather = get_weather(q)
        print(get_weather_desc(weather))
        save_excel(weather)
